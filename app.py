import os
import re
from datetime import datetime
from io import BytesIO
from flask import Flask, render_template_string, request, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
import openpyxl

app = Flask(__name__)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'diario_producao_estavel.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- MODELOS DO BANCO DE DADOS ---
class Turma(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    escola = db.Column(db.String(150), nullable=False)
    nome_turma = db.Column(db.String(50), nullable=False)
    disciplina = db.Column(db.String(100), nullable=False)
    alunos = db.relationship('Aluno', backref='turma', lazy=True, cascade="all, delete-orphan")

class Aluno(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    num_chamada = db.Column(db.String(10), nullable=True)
    matricula = db.Column(db.String(50), nullable=True)
    nome = db.Column(db.String(200), nullable=False)
    situacao = db.Column(db.String(50), nullable=True)
    nota1 = db.Column(db.Float, default=0.0)
    nota2 = db.Column(db.Float, default=0.0)
    turma_id = db.Column(db.Integer, db.ForeignKey('turma.id'), nullable=False)
    presencas = db.relationship('Presenca', backref='aluno', lazy=True, cascade="all, delete-orphan")

class Presenca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.String(10), nullable=False)
    status = db.Column(db.String(1), nullable=False)
    aluno_id = db.Column(db.Integer, db.ForeignKey('aluno.id'), nullable=False)

with app.app_context():
    db.create_all()

# --- INTERFACE HTML ---
HTML_COMPLETO = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Portal do Docente - CIEP</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
    <style>
        body { background-color: #f4f6f9; font-family: system-ui, sans-serif; }
        .navbar-custom { background-color: #1a365d; color: white; }
        .card-custom { border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05); border: none; }
        .table th { background-color: #1a365d !important; color: white !important; padding: 12px; text-align: center; }
        .table td { vertical-align: middle; padding: 10px; }
        .input-nota { width: 75px; text-align: center; font-weight: bold; }
    </style>
</head>
<body>
<nav class="navbar navbar-custom p-3 mb-4">
    <div class="container d-flex justify-content-between">
        <span class="navbar-brand mb-0 h1 text-white">🍎 Diário de Frequência Inteligente</span>
        <a href="/" class="btn btn-sm btn-outline-light fw-bold">🏠 Voltar ao Menu</a>
    </div>
</nav>
<div class="container">
    {% if tela == 'inicial' %}
    <div class="row">
        <div class="col-md-5 mb-4">
            <div class="card card-custom p-4 bg-white">
                <h5 class="fw-bold text-primary mb-3">📂 Carregar Novo Diário (CSV)</h5>
                <form action="/carregar-csv" method="POST" enctype="multipart/form-data">
                    <div class="mb-3">
                        <label class="form-label small fw-bold">Selecione o Arquivo CSV</label>
                        <input type="file" name="arquivo_csv" class="form-control form-control-sm" accept=".csv" required>
                    </div>
                    <button type="submit" class="btn btn-primary btn-sm w-100 fw-bold">🔄 Criar Diário na Nuvem</button>
                </form>
            </div>
        </div>
        <div class="col-md-7">
            <div class="card card-custom p-4 bg-white">
                <h5 class="fw-bold text-success mb-3">📚 Suas Turmas Salvas na Nuvem</h5>
                {% if not turmas %}
                    <p class="text-muted small">Nenhuma turma cadastrada ainda.</p>
                {% else %}
                    <div class="list-group">
                        {% for t in turmas %}
                            <div class="list-group-item d-flex justify-content-between align-items-center mb-2 rounded border">
                                <div>
                                    <h6 class="fw-bold m-0">Turma {{ t.nome_turma }} - {{ t.disciplina }}</h6>
                                    <small class="text-muted">{{ t.escola }}</small>
                                </div>
                                <div>
                                    <a href="/chamada/{{ t.id }}" class="btn btn-sm btn-success fw-bold">📅 Caderneta</a>
                                    <a href="/excluir-turma/{{ t.id }}" class="btn btn-sm btn-outline-danger">🗑️</a>
                                </div>
                            </div>
                        {% endfor %}
                    </div>
                {% endif %}
            </div>
        </div>
    </div>
    {% elif tela == 'chamada' %}
    <div class="card card-custom p-4 bg-white mb-4">
        <div class="d-flex justify-content-between align-items-center mb-3 pb-2 border-bottom">
            <div>
                <h4 class="fw-bold m-0 text-dark">📋 Fechamento de Notas e Frequência</h4>
                <small class="text-muted">{{ turma.escola }} | {{ turma.disciplina }} | Turma: {{ turma.nome_turma }}</small>
                <div class="mt-2"><span class="badge bg-primary fs-6">📅 Data Selecionada: 16/06/2026</span></div>
            </div>
            <a href="/baixar-excel/{{ turma.id }}" class="btn btn-success btn-sm fw-bold px-4 shadow-sm">📥 Exportar Planilha de Impressão</a>
        </div>
        <form action="/salvar-dados/{{ turma.id }}" method="POST">
            <input type="hidden" name="data_chamada" value="{{ data_atual }}">
            <div class="table-responsive">
                <table class="table table-striped table-bordered align-middle m-0 table-sm">
                    <thead>
                        <tr>
                            <th style="width: 50px;">Nº</th>
                            <th>Nome Completo do Aluno</th>
                            <th style="width: 80px;">Faltas</th>
                            <th style="width: 95px;">Nota 1</th>
                            <th style="width: 95px;">Nota 2</th>
                            <th style="width: 95px;">Média</th>
                            <th style="width: 130px;">Chamada de Hoje (16/06)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for aluno in alunos_info %}
                        <tr>
                            <td class="text-center text-muted small">{{ aluno.num_chamada }}</td>
                            <td><strong class="text-dark">{{ aluno.nome }}</strong><br><small class="text-muted">{{ aluno.matricula }}</small></td>
                            <td class="text-center text-danger fw-bold">{{ aluno.total_faltas }}</td>
                            <td class="text-center">
                                <input type="number" step="0.1" min="0" max="10" name="nota1_{{ aluno.id }}" class="form-control form-control-sm input-nota text-primary" value="{{ aluno.nota1 }}">
                            </td>
                            <td class="text-center">
                                <input type="number" step="0.1" min="0" max="10" name="nota2_{{ aluno.id }}" class="form-control form-control-sm input-nota text-primary" value="{{ aluno.nota2 }}">
                            </td>
                            <td class="text-center fw-bold {% if aluno.media >= 6.0 %}text-success{% else %}text-danger{% endif %}">
                                {{ "%.1f"|format(aluno.media) }}
                            </td>
                            <td class="text-center">
                                <div class="btn-group">
                                    <input type="radio" class="btn-check" name="status_{{ aluno.id }}" id="p_{{ aluno.id }}" value="P" {% if aluno.status_hoje == 'P' %}checked{% endif %}>
                                    <label class="btn btn-xs btn-outline-success px-2 fw-bold" for="p_{{ aluno.id }}">P</label>
                                    <input type="radio" class="btn-check" name="status_{{ aluno.id }}" id="f_{{ aluno.id }}" value="F" {% if aluno.status_hoje == 'F' %}checked{% endif %}>
                                    <label class="btn btn-xs btn-outline-danger px-2 fw-bold" for="f_{{ aluno.id }}">F</label>
                                </div>
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            <button type="submit" class="btn btn-primary fw-bold px-5 mt-3 shadow">💾 Gravar Chamada e Notas (Dia 16/06)</button>
        </form>
    </div>
    {% endif %}
</div>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_COMPLETO, tela='inicial', turmas=Turma.query.all())

@app.route('/carregar-csv', methods=['POST'])
def carregar_csv():
    file = request.files.get('arquivo_csv')
    if file:
        try:
            texto_bruto = file.read().decode('utf-8', errors='ignore')
            linhas = texto_bruto.replace('\r', '\n').split('\n')
            
            escola_auto = "CIEP DOUTOR ULYSSES GUIMARAES"
            turma_auto = "1017"
            disciplina_auto = "DIÁRIO DE CLASSE"
            
            for l in linhas[:5]:
                if l and "CIEP" in l.upper() and "ANDRE" in l.upper():
                    partes = [p.replace('"', '').strip() for p in re.split(r'[;,]', l) if p.strip()]
                    if len(partes) >= 3:
                        escola_auto = partes[0]
                        disciplina_auto = partes[-1]

            nova_turma = Turma(escola=escola_auto, nome_turma=turma_auto, disciplina=disciplina_auto)
            db.session.add(nova_turma)
            db.session.commit()

            contador_chamada = 1

            for linha in linhas:
                if not linha or not isinstance(linha, str):
                    continue
                linha_up = linha.upper()
                
                if "NUM_CHAMADA" in linha_up or "ANDRE CAMARGO" in linha_up or "CANCELADO" in linha_up:
                    continue
                
                if "MATRICULADO" in linha_up:
                    linha_limpa = linha.replace('"', '')
                    partes = [p.strip() for p in re.split(r'[;,]', linha_limpa) if p.strip()]
                    
                    matricula = ""
                    textos = []
                    
                    for p in partes:
                        if p.isdigit() and len(p) >= 10:
                            matricula = p
                        elif len(p) > 5 and not any(c.isdigit() for c in p) and "MATRICULADO" not in p.upper() and "TRIMESTRE" not in p.upper():
                            textos.append(p.upper())
                    
                    name_found = textos[0] if textos else ""
                    
                    if name_found and matricula:
                        db.session.add(Aluno(
                            num_chamada=str(contador_chamada),
                            matricula=str(matricula),
                            nome=str(name_found),
                            situacao="MATRICULADO",
                            turma_id=nova_turma.id
                        ))
                        contador_chamada += 1
                        
            db.session.commit()
        except Exception as e:
            print(f"Erro na importação: {e}")
    return redirect(url_for('index'))

@app.route('/chamada/<int:turma_id>')
def chamada(turma_id):
    turma = Turma.query.get_or_404(turma_id)
    data_filtro = "2026-06-16"
    alunos = Aluno.query.filter_by(turma_id=turma.id).all()
    alunos_ordenados = sorted(alunos, key=lambda x: int(x.num_chamada) if (x.num_chamada and str(x.num_chamada).isdigit()) else 99)
    
    alunos_info = []
    for a in alunos_ordenados:
        reg = Presenca.query.filter_by(aluno_id=a.id, data=data_filtro).first()
        status_hoje = reg.status if reg else 'P'
        total_faltas = Presenca.query.filter_by(aluno_id=a.id, status='F').count()
        media = (float(a.nota1 or 0) + float(a.nota2 or 0)) / 2
        
        alunos_info.append({
            "id": a.id, "num_chamada": a.num_chamada, "matricula": a.matricula,
            "nome": a.nome, "situacao": a.situacao, "status_hoje": status_hoje,
            "total_faltas": total_faltas, "nota1": a.nota1, "nota2": a.nota2, "media": media
        })
    return render_template_string(HTML_COMPLETO, tela='chamada', turma=turma, alunos_info=alunos_info, data_atual=data_filtro)

@app.route('/salvar-dados/<int:turma_id>', methods=['POST'])
def salvar_dados(turma_id):
    data_chamada = request.form.get('data_chamada', '2026-06-16')
    alunos = Aluno.query.filter_by(turma_id=turma_id).all()
    
    for a in alunos:
        n1 = request.form.get(f'nota1_{a.id}')
        n2 = request.form.get(f'nota2_{a.id}')
        try: a.nota1 = float(n1) if n1 else 0.0
        except: a.nota1 = 0.0
        try: a.nota2 = float(n2) if n2 else 0.0
        except: a.nota2 = 0.0
        
        status = request.form.get(f'status_{a.id}', 'P')
        
        aluno_id_num = int(a.id)
        reg = db.session.query(Presenca).filter(Presenca.aluno_id == aluno_id_num, Presenca.data == str(data_chamada)).first()
        if reg:
            reg.status = str(status)
        else:
            db.session.add(Presenca(data=str(data_chamada), status=str(status), aluno_id=aluno_id_num))
            
    db.session.commit()
    return redirect(url_for('chamada', turma_id=turma_id))

@app.route('/excluir-turma/<int:turma_id>')
def excluir_turma(turma_id):
    turma = Turma.query.get(turma_id)
    if r:= turma:
        db.session.delete(r)
        db.session.commit()
    return redirect(url_for('index'))

@app.route('/baixar-excel/<int:turma_id>')
def baixar_excel(turma_id):
    try:
        turma = Turma.query.get_or_404(turma_id)
        alunos = Aluno.query.filter_by(turma_id=turma.id).all()
        alunos_ordenados = sorted(alunos, key=lambda x: int(x.num_chamada) if (x.num_chamada and str(x.num_chamada).isdigit()) else 99)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Turma {turma.nome_turma}"
        
        ws.views.sheetView[0].showGridLines = True
        
        headers = ["Nº", "Matrícula", "Nome Completo do Aluno", "Total Faltas", "Nota 1", "Nota 2", "Média Final"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        
        for idx, a in enumerate(alunos_ordenados, 2):
            total_faltas = Presenca.query.filter_by(aluno_id=a.id, status='F').count()
            media = (float(a.nota1 or 0) + float(a.nota2 or 0)) / 2
            
            ws.cell(row=idx, column=1, value=str(a.num_chamada) if a.num_chamada else str(idx-1))
            ws.cell(row=idx, column=2, value=str(a.matricula) if a.matricula else "")
            ws.cell(row=idx, column=3, value=str(a.nome) if a.nome else "")
            ws.cell(row=idx, column=4, value=str(total_faltas))
            ws.cell(row=idx, column=5, value=str(a.nota1))
            ws.cell(row=idx, column=6, value=str(a.nota2))
            ws.cell(row=idx, column=7, value=str(media))
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f"Diario_Fechamento_Turma_{turma.nome_turma}.xlsx")
    except Exception as e:
        return f"Erro ao gerar planilha: {str(e)}", 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)
